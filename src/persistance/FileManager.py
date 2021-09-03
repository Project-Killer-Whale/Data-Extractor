from src. entities.BeachInformation import BeachInformation
from src.persistance.JSONFormatter import JSONFormatter
from src.entities.BeachCode import *
from src.entities.Town import *

import openpyxl
import unidecode
import codecs
import csv
import os

class FileManager():

    def save_data(ou_file):
        file = open(ou_file.path, "w")
        file.write(ou_file.data)
        file.close()

    def check_if_exists(path):
        return os.path.exists(path)

    def create_dir(parent_path, name):
        path = parent_path + name
        os.mkdir(path) 
        return path

    def read_xlsx_file(file_path, first_row, last_row, columns):
        data = []

        file = openpyxl.load_workbook(file_path)
        sheet = file.active

        for row in range(first_row, last_row + 1):

            cpro = sheet.cell(row=row, column=columns[0]).value
            cmun = sheet.cell(row=row, column=columns[1]).value
            name = sheet.cell(row=row, column=columns[2]).value

            data.append(Town(str(cpro) + str(cmun), unidecode.unidecode(name)))

        return data

    def read_csv_file(file_path, separator):
        data = []

        with open(file_path, 'r', encoding="latin1", errors="replace") as csv_file:
            file = csv.reader(csv_file, delimiter = separator)
        
            for line in file:
                data.append(BeachCode(line[0],line[1],line[4],line[5],line[2],line[3],line[6], line[7]))

        return data

    def read_txt_file(path):
        result = ""
        
        with open(path, 'r') as file:
            result = file.read()
        
        return result
    
    def read_json_file(path):
        data = []
        json_result = []

        with open(path, 'r') as file:
            json_result = JSONFormatter.convert_json_to_object(file.read())
            for item in json_result:
                beach = BeachInformation()
                beach.assign_object_dict(item)
                data.append(beach)

        return data